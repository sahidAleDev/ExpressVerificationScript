from abc import ABC, abstractmethod
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import csv
import json


class DataReader(ABC):
    @abstractmethod
    def read_data(self, filename):
        pass


class CSVReader(DataReader):
    def read_data(self, filename):
        with open(filename) as file:
            csv_data = csv.reader(file)
            next(csv_data)
            next(csv_data)
            next(csv_data)

            client_names = []
            aval_names = []
            for row in csv_data:
                full_name_client = row[2] + " " + row[3] + " " + row[4]
                full_name_aval = row[408] + " " + row[409] + " " + row[410]
                id_prestamo = row[12]
                week = int(row[15])
                year = int(row[16])
                balance = row[404].strip()
                client_names.append([id_prestamo, full_name_client, week, year, balance])
                aval_names.append([id_prestamo, full_name_aval])

        return client_names, aval_names


class JsonReader(DataReader):
    def read_data(self, filename):
        with open(filename) as file:
            data = json.load(file)

        rows = []
        for person in data:
            id_prestamo = person['prestamoId']
            full_name = person['nombres'] + " " + person['apellidoPaterno'] + " " + person['apellidoMaterno']
            week = person['semana']
            year = person['anio']
            rows.append([id_prestamo, full_name, week, year])

        return rows


class ListComparer:
    def compare_lists(self, list1, list2):
        result = [[item, item in list2] for item in list1]
        return result


class BalanceUpdater: # Agregar el saldo a los datos de 
    def add_balance_in_odoodata(self, list_excel_crudo, list_odoo):
        id_index = 0
        name_index = 1
        last_element_index = -1

        for sublist in list_odoo:
            sublist.append("X")

        for sublist1 in list_excel_crudo:
            for sublist2 in list_odoo:
                if (sublist1[id_index] == sublist2[id_index]) and (sublist1[name_index] == sublist2[name_index]) and (sublist1[last_element_index] != " "):
                    sublist2[4] = sublist1[last_element_index]


class ExcelFileCreator:
    def __init__(self):
        self.workbook = Workbook()

    def create_excel(self, file_name, data, source_of_comparison):
        sheet = self.workbook.active
        total_records, match, no_match = self.get_matches(data)

        sheet['A1'] = 'IdPrestamo'
        sheet['B1'] = 'Nombre'
        sheet['C1'] = 'Semana'
        sheet['D1'] = 'Año'
        sheet['E1'] = 'Saldo'
        sheet['F1'] = '¿Se encuentra en %s?' % source_of_comparison

        sheet['G1'] = "Total de personas"
        sheet['H1'] = "Coincidencias"
        sheet['I1'] = "No Coincidencias"
        sheet['G2'] = total_records
        sheet['H2'] = match
        sheet['I2'] = no_match

        for i, dato in enumerate(data, start=2):
            sheet[f'A{i}'] = dato[0][0]
            sheet[f'B{i}'] = dato[0][1]
            sheet[f'C{i}'] = dato[0][2]
            sheet[f'D{i}'] = dato[0][3]
            sheet[f'E{i}'] = dato[0][4]
            sheet[f'F{i}'] = dato[1]
            color = ""

            if dato[0][4] == '$-' or dato[0][4] == "X":
                color = "F48498"  # Ejemplo: rojo en formato RGB
            else:
                color = "ACD8AA"

            # Crea un objeto PatternFill con el color de relleno
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Aplica el relleno a una celda específica
            sheet[f'E{i}'].fill = fill

        full_file_name = file_name + ".xlsx"
        self.workbook.save(full_file_name)
        self.workbook.close()
        print("Archivo %s creado con éxito\n" % full_file_name)

    def get_matches(self, data):
        match = 0
        no_match = 0
        for item in data:
            if item[1]:
                match += 1
            else:
                no_match += 1

        return len(data), match, no_match


def main():
    #* Lectores de Archivos
    excel_csv_reader = CSVReader()
    odoo_json_reader = JsonReader()

    csv_semana_cruda_client, csv_semana_cruda_aval = excel_csv_reader.read_data("BD DINERO XPRESS 2023. CRUDA SEMANA 25 CSV.csv")
    json_odoo_client = odoo_json_reader.read_data("prestamos_v2_semana_23_anio_2023.json")

    #* Agregar el saldo a los datos del JSON (No viene en él)
    balance_updater = BalanceUpdater()
    balance_updater.add_balance_in_odoodata(csv_semana_cruda_client, json_odoo_client)

    #* Comparacion de las listas de los datos
    excelclients_with_odoo = ListComparer().compare_lists(csv_semana_cruda_client, json_odoo_client)
    odooclients_with_excel = ListComparer().compare_lists(json_odoo_client, csv_semana_cruda_client)

    #* Creacion de los archivos de Excel
    workbook = Workbook()
    excel_creator = ExcelFileCreator()
    excel_creator.create_excel("excel_clients_in_odoo", excelclients_with_odoo, "Odoo")
    workbook.close()

    workbook = Workbook()
    excel_creator = ExcelFileCreator()
    excel_creator.create_excel("odoo_clients_in_excel", odooclients_with_excel, "Excel")
    workbook.close()


if __name__ == "__main__":
    main()