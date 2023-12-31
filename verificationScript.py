from abc import ABC, abstractmethod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import csv
import json


class DataReader(ABC):
    @abstractmethod
    def read_data(self, filename):
        pass


class CSVReader(DataReader):
    def read_data(self, filename):
        with open(filename) as file:
            csv_data = csv.reader(file)
            next(csv_data)
            next(csv_data)
            next(csv_data)

            loans = []
            for loan in csv_data:
                loan_id = loan[0]
                client_full_name = loan[2] + " " + loan[3] + " " + loan[4]
                aval_full_name = loan[408] + " " + loan[409] + " " + loan[410]
                agent = loan[12]
                management = loan[13]
                week = int(loan[15])
                year = int(loan[16])
                balance = loan[404].strip()
                loans.append([loan_id, client_full_name, week, year, aval_full_name, agent, management, balance])

        return loans


class JsonReader(DataReader):
    def read_data(self, filename):
        with open(filename) as file:
            data = json.load(file)

        loans = []

        for loan in data:
            loan_id = loan['prestamoId']
            client_full_name = loan['nombres'] + " " + loan['apellidoPaterno'] + " " + loan['apellidoMaterno']
            aval_full_name = loan['nombresAval'] + " " + loan['apellidoPaternoAval'] + " " + loan['apellidoMaternoAval']
            week = loan['semana']
            year = loan['anio']
            agent = loan['agente']
            management = loan['gerencia']
            loans.append([loan_id, client_full_name, week, year, aval_full_name, agent, management])

        return loans


class ListComparer:
    def compare_lists(self, list1, list2):
        # result = [[item, item in list2] for item in list1]
        # return result
        result = []
        for item1 in list1:
            matching_item = False
            for item2 in list2:
                if item1[0] == item2[0]:
                    matching_item = True
                    break
            result.append([item1, matching_item])
        return result


class BalanceUpdater:
    def add_balance_in_odoodata(self, list_excel_crudo, list_odoo):
        id_index = 0
        name_client_index = 1
        name_aval_index = 4
        last_element_index = -1

        for sublist in list_odoo:
            sublist.append("X")

        for sublist1 in list_excel_crudo:
            for sublist2 in list_odoo:
                if (sublist1[id_index] == sublist2[id_index] and sublist1[last_element_index] != " "):
                    sublist2[7] = sublist1[last_element_index]


class ExcelFileCreator:
    def __init__(self):
        self.workbook = None

    def create_excel(self, file_name, data, source_of_comparison):
        if self.workbook is None:
            self.workbook = Workbook()

        sheet = self.workbook.active
        total_records, match, no_match = self.get_matches(data)

        font = Font(bold=True)
        sheet['A1'] = 'IdPrestamo'
        sheet['B1'] = 'Nombre Cliente'
        sheet['C1'] = 'Semana'
        sheet['D1'] = 'Año'
        sheet['E1'] = 'Nombre del Aval'
        sheet['F1'] = 'Agente'
        sheet['G1'] = 'Gerencia'


        sheet['H1'] = 'Saldo'
        sheet['I1'] = 'Cant. Prestamos del Cliente'
        sheet['J1'] = '¿Se encuentra en %s?' % source_of_comparison

        sheet['L1'] = "Total de prestamos"
        sheet['M1'] = "Coincidencias en %s" % source_of_comparison
        sheet['N1'] = "No Coincidencias"
        sheet['L2'] = total_records
        sheet['M2'] = match
        sheet['N2'] = no_match

        sheet['A1'].font = sheet['B1'].font = sheet['C1'].font = sheet['D1'].font = sheet['E1'].font = sheet['F1'].font = sheet['G1'].font = sheet['H1'].font = sheet['I1'].font = sheet['J1'].font = font
        
        for i, dato in enumerate(data, start=2):
            sheet[f'A{i}'] = dato[0][0]
            sheet[f'B{i}'] = dato[0][1]
            sheet[f'C{i}'] = dato[0][2]
            sheet[f'D{i}'] = dato[0][3]
            sheet[f'E{i}'] = dato[0][4]
            sheet[f'F{i}'] = dato[0][5]
            sheet[f'G{i}'] = dato[0][6]
            sheet[f'H{i}'] = dato[0][7]
            color = ""

            if dato[0][7] == '$-' or dato[0][7] == "X": # Estado segun el saldo
                color = "F48498"  
            else:
                color = "ACD8AA"

            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sheet[f'H{i}'].fill = fill

            sheet[f'I{i}'] = dato[2]
            sheet[f'J{i}'] = dato[1]

        full_file_name = file_name + ".xlsx"
        self.workbook.save(full_file_name)
        self.workbook.close()
        self.workbook = None
        print("Archivo %s creado con éxito\n" % full_file_name)

    def get_matches(self, data):
        match = 0
        no_match = 0
        for item in data:
            if item[1]:
                match += 1
            else:
                no_match += 1

        return len(data), match, no_match


class LoanCalculator:
    def calculate_loan_amount(self, loan_data):
        count_dict = {}
        for item in loan_data:
            full_name = item[0][1]
            if full_name in count_dict:
                count_dict[full_name] += 1
            else:
                count_dict[full_name] = 1

        for item in loan_data:
            full_name = item[0][1]
            item.append(count_dict[full_name])


class Application:
    def __init__(self, excel_csv_reader, odoo_json_reader, balance_updater, list_comparer, excel_creator, loan_calculator):
        self.excel_csv_reader = excel_csv_reader
        self.odoo_json_reader = odoo_json_reader
        self.balance_updater = balance_updater
        self.list_comparer = list_comparer
        self.excel_creator = excel_creator
        self.loan_calculator = loan_calculator

    def run(self):
        csv_sem_cruda_loan_data = self.excel_csv_reader.read_data("BD DINERO XPRESS 2023. CRUDA SEMANA 27 CSV.csv")
        json_odoo_loan_data = self.odoo_json_reader.read_data("prestamos_semana28.json")

        self.balance_updater.add_balance_in_odoodata(csv_sem_cruda_loan_data, json_odoo_loan_data)

        excelclients_with_odoo = self.list_comparer.compare_lists(csv_sem_cruda_loan_data, json_odoo_loan_data)
        odooclients_with_excel = self.list_comparer.compare_lists(json_odoo_loan_data, csv_sem_cruda_loan_data)

        self.loan_calculator.calculate_loan_amount(excelclients_with_odoo)
        self.loan_calculator.calculate_loan_amount(odooclients_with_excel)

        self.excel_creator.create_excel("excel_clients_in_odoo", excelclients_with_odoo, "Odoo")
        self.excel_creator.create_excel("odoo_clients_in_excel", odooclients_with_excel, "Excel")


def main():
    excel_csv_reader = CSVReader()
    odoo_json_reader = JsonReader()
    balance_updater = BalanceUpdater()
    list_comparer = ListComparer()
    excel_creator = ExcelFileCreator()
    loan_calculator = LoanCalculator()

    app = Application(
        excel_csv_reader,
        odoo_json_reader,
        balance_updater,
        list_comparer,
        excel_creator,
        loan_calculator
    )
    app.run()


if __name__ == "__main__":
    main()
